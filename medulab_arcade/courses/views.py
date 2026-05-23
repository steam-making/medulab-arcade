import io
import json
import os
import sys
import re
import hashlib
import secrets
import qrcode
from html import unescape
from collections import Counter
import tempfile
import zipfile
from datetime import timedelta
from types import SimpleNamespace
from xml.etree import ElementTree as ET
import openpyxl
from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, Http404
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.db import transaction
from django.core.files.base import ContentFile
from django.core.files.images import get_image_dimensions
from arcade.badge_service import (
    evaluate_homework_badges,
    evaluate_mission_badges,
    evaluate_program_badges,
    get_program_completion_badge,
    get_recent_user_badges,
)
from .models import (
    AnswerZipImportBatch,
    LearningProgram,
    Chapter,
    Item,
    LearningEnrollment,
    UserProgress,
    ProgramType,
    HomeworkAssignment,
    HomeworkAttachment,
    HomeworkSubmission,
    OlympiadAnswerSubmission,
    OlympiadAnswerExample,
    OlympiadSubQuestion,
    OlympiadSubAnswer,
)
from .forms import (
    AnswerZipImportForm,
    ChapterForm,
    CourseForm,
    ProgramTypeForm,
    ItemForm,
    HomeworkForm,
    HomeworkSubmissionForm,
    HomeworkSubmissionReviewForm,
    OlympiadAnswerSubmissionForm,
)
from django.db.models import Count, Q

PPT_EXAM_DURATION_SECONDS = 60 * 60
PPT_SLIDE_XML_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
ANSWER_ZIP_ITEM_KEY_PREFIX = "answer_zip"
ANSWER_ZIP_MAX_FILES = 500
ANSWER_ZIP_MAX_TOTAL_BYTES = 200 * 1024 * 1024
ANSWER_ZIP_MAX_TOTAL_TEXT_BYTES = 5 * 1024 * 1024
ANSWER_ZIP_MAX_TEXT_FILE_BYTES = 512 * 1024
ANSWER_ZIP_MAX_PPTX_FILE_BYTES = 25 * 1024 * 1024
ANSWER_ZIP_MAX_UPLOAD_BYTES = 200 * 1024 * 1024
ANSWER_ZIP_TEXT_EXTENSIONS = {".txt", ".py", ".md", ".html", ".htm", ".json", ".csv", ".xml", ".yaml", ".yml", ".js", ".css"}
ANSWER_ZIP_OLYMPIAD_RULE = "olympiad_answer_examples_by_chapter_item"
ANSWER_ZIP_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}
OLYMPIAD_EXAMPLE_MAX_IMAGE_BYTES = 10 * 1024 * 1024
OLYMPIAD_EXAMPLE_MAX_IMAGE_PIXELS = 40_000_000


class AnswerZipImportError(ValueError):
    pass


def natural_sort_key(value):
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", value or "")]


def normalize_answer_zip_path(raw_name):
    name = (raw_name or "").strip()
    if not name or name.endswith("/"):
        return None
    if "\\" in name or name.startswith("/") or re.match(r"^[A-Za-z]:", name):
        raise AnswerZipImportError(f"안전하지 않은 ZIP 경로입니다: {raw_name}")

    parts = [part for part in name.split("/") if part]
    if len(parts) < 2:
        raise AnswerZipImportError(f"A-rule 위반: 최상위 파일은 허용하지 않습니다 ({raw_name}).")
    if any(part in {".", ".."} for part in parts):
        raise AnswerZipImportError(f"상위 경로 참조는 허용하지 않습니다: {raw_name}")
    if parts[0] == "__MACOSX" or parts[-1].startswith("._"):
        return None
    return parts


def decode_answer_zip_filename(info):
    filename = info.filename or ""
    if info.flag_bits & 0x800:
        return filename

    try:
        repaired = filename.encode("cp437").decode("cp949")
        if repaired:
            return repaired
    except UnicodeError:
        pass

    return filename


def slug_for_key(value):
    slug = re.sub(r"[^0-9A-Za-z가-힣]+", "_", value or "").strip("_").lower()
    return slug or "item"


def make_answer_zip_item_key(chapter_title, item_path):
    base = f"{ANSWER_ZIP_ITEM_KEY_PREFIX}_{slug_for_key(chapter_title)}_{slug_for_key(item_path)}"
    if len(base) <= 100:
        return base
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:8]
    return base[:91].rstrip("_") + f"_{digest}"


def title_from_zip_path(relative_path):
    basename = relative_path.split("/")[-1]
    title = re.sub(r"\.[^.]+$", "", basename)
    title = re.sub(r"^[\d\s._-]+", "", title).strip()
    return title or basename


def is_text_like_answer_file(relative_path):
    _, extension = os.path.splitext(relative_path or "")
    return extension.lower() in ANSWER_ZIP_TEXT_EXTENSIONS


def is_pptx_answer_file(relative_path):
    _, extension = os.path.splitext(relative_path or "")
    return extension.lower() == ".pptx"


def is_image_answer_file(relative_path):
    _, extension = os.path.splitext(relative_path or "")
    return extension.lower() in ANSWER_ZIP_IMAGE_EXTENSIONS


def validate_olympiad_example_image(image):
    if image.size > OLYMPIAD_EXAMPLE_MAX_IMAGE_BYTES:
        return "예시답안 이미지는 10MB 이하만 등록할 수 있습니다."
    if not is_image_answer_file(image.name):
        return "예시답안 이미지는 jpg, jpeg, png, gif, webp, bmp 파일만 등록할 수 있습니다."

    try:
        image.seek(0)
        width, height = get_image_dimensions(image)
    except Exception:
        return "올바른 이미지 파일만 등록할 수 있습니다."
    finally:
        image.seek(0)

    if not width or not height:
        return "올바른 이미지 파일만 등록할 수 있습니다."
    if width * height > OLYMPIAD_EXAMPLE_MAX_IMAGE_PIXELS:
        return "이미지 해상도가 너무 큽니다. 더 작은 이미지로 등록해 주세요."

    return None


def format_olympiad_image_size(size):
    if not size:
        return "파일 크기를 확인할 수 없습니다"
    if size >= 1024 * 1024:
        return f"{size / (1024 * 1024):.1f}MB"
    if size >= 1024:
        return f"{size / 1024:.1f}KB"
    return f"{size}B"


def get_olympiad_image_dimensions(image):
    if not image:
        return None
    try:
        image.seek(0)
        width, height = get_image_dimensions(image)
    except Exception:
        return None
    finally:
        try:
            image.seek(0)
        except Exception:
            pass
    if not width or not height:
        return None
    return width, height


def summarize_olympiad_hint(hint):
    text = re.sub(r"<[^>]+>", " ", unescape(hint or ""))
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= 90:
        return text
    return text[:90].rstrip() + "..."


def calculate_olympiad_ai_score(submission):
    answer_text = (submission.edited_text or submission.ocr_text or "").strip()
    image = submission.answer_image
    image_size = int(getattr(image, "size", 0) or 0)
    dimensions = get_olympiad_image_dimensions(image)

    score = 60
    if answer_text:
        score += 10
        if len(answer_text) >= 80:
            score += 8
        elif len(answer_text) >= 35:
            score += 5
    if dimensions:
        width, height = dimensions
        pixels = width * height
        if pixels >= 800 * 600:
            score += 8
        elif pixels >= 500 * 300:
            score += 5
        if min(width, height) >= 280:
            score += 4
    if image_size >= 100 * 1024:
        score += 5
    if 0 < image_size <= 8 * 1024 * 1024:
        score += 4
    if not answer_text:
        score = min(score, 82)
    return max(45, min(score, 95))


def parse_olympiad_feedback(feedback):
    section_titles = {"AI 평가점수", "사진 확인", "답안 보완 포인트", "다음 제출 팁"}
    sections = []
    current = None

    for raw_line in (feedback or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line in section_titles:
            current = {"title": line, "items": [], "kind": "default", "score": None}
            if line == "AI 평가점수":
                current["kind"] = "score"
            elif line == "답안 보완 포인트":
                current["kind"] = "improvement"
            elif line == "다음 제출 팁":
                current["kind"] = "tip"
            else:
                current["kind"] = "photo"
            sections.append(current)
            continue
        if current is None:
            current = {"title": "제출 피드백", "items": [], "kind": "default", "score": None}
            sections.append(current)
        cleaned = line[2:].strip() if line.startswith("- ") else line
        score_match = re.search(r"(\d{1,3})\s*/\s*100", cleaned)
        if score_match:
            current["score"] = max(0, min(int(score_match.group(1)), 100))
            if current["kind"] == "score":
                continue
        current["items"].append(cleaned)

    return sections


def build_olympiad_submission_feedback(submission, item):
    image = submission.answer_image
    image_size = format_olympiad_image_size(getattr(image, "size", 0))
    dimensions = get_olympiad_image_dimensions(image)
    dimension_text = f"{dimensions[0]}x{dimensions[1]}px" if dimensions else "해상도 확인 불가"
    answer_text = (submission.edited_text or submission.ocr_text or "").strip()
    hint_summary = summarize_olympiad_hint(item.hint)
    ai_score = calculate_olympiad_ai_score(submission)

    feedback_lines = [
        "AI 평가점수",
        f"- {ai_score}/100점",
        "- 사진 선명도, 답안 보충 텍스트, 문제 조건 확인 가능성을 기준으로 자동 산정했습니다.",
        "",
        "사진 확인",
        f"- 업로드된 답안 사진은 {image_size}, {dimension_text}로 확인되었습니다.",
        "- 글자 가장자리, 문제 번호, 풀이 순서가 한 장 안에서 선명하게 보이는지 다시 확인해 보세요.",
        "",
        "답안 보완 포인트",
    ]
    if answer_text:
        feedback_lines.append("- 적어 준 답안 내용을 기준으로 조건, 풀이 과정, 결론이 모두 드러나는지 점검해 보세요.")
    else:
        feedback_lines.append("- 답안 사진만으로도 제출은 완료되었습니다. 다음에는 핵심 풀이를 글로 한 번 더 적으면 스스로 빠진 조건을 찾기 쉽습니다.")
    if hint_summary:
        feedback_lines.append(f"- 문제 힌트의 핵심인 '{hint_summary}'와 연결되는 설명을 답안에 포함했는지 확인해 보세요.")
    feedback_lines.extend([
        "- 답을 고른 이유와 계산/분류 기준을 번호나 표로 정리하면 채점자가 사고 과정을 따라가기 좋습니다.",
        "",
        "다음 제출 팁",
        "- 그림자 없이 밝은 곳에서 종이를 화면 가득 맞추고, 업로드 전 흔들림과 잘림이 없는지 확인하세요.",
        "- 수정할 내용이 보이면 사진을 다시 찍거나 답안 내용 칸에 빠진 문장을 보충해 주세요.",
    ])
    return "\n".join(feedback_lines)


def extract_first_number(value):
    match = re.search(r"\d+", value or "")
    return int(match.group()) if match else None


def build_olympiad_answer_template_zip(program):
    buffer = io.BytesIO()
    chapters = program.chapters.prefetch_related("items").order_by("number")
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "README.txt",
            "각 챕터/문제 폴더 안에 예시답안 이미지 파일을 넣은 뒤 다시 업로드하세요.\n"
            "예: 제1회 기출문제/문제 2/01.jpg\n"
            "지원 형식: jpg, jpeg, png, gif, webp, bmp\n",
        )
        for chapter in chapters:
            chapter_folder = f"{chapter.title}/"
            archive.writestr(chapter_folder, "")
            for item in chapter.items.filter(item_type="olympiad").order_by("number"):
                problem_folder = f"{chapter.title}/{item.title}/"
                archive.writestr(problem_folder, "")
                archive.writestr(
                    f"{problem_folder}README.txt",
                    f"{chapter.title} - {item.title} 예시답안 이미지를 이 폴더에 넣으세요.\n"
                    "파일명 예: 01.jpg, 02.png\n",
                )
    buffer.seek(0)
    return buffer


def parse_olympiad_answer_examples_zip(program, uploaded_file):
    uploaded_file.seek(0)
    matched_items = {}
    unmatched_files = []
    file_count = 0
    total_bytes = 0

    items_by_chapter_and_number = {
        (item.chapter.number, item.number): item
        for item in Item.objects.filter(chapter__program=program, item_type="olympiad").select_related("chapter")
    }

    try:
        with zipfile.ZipFile(uploaded_file) as archive:
            infos = archive.infolist()
            if not infos:
                raise AnswerZipImportError("비어 있는 ZIP 파일입니다.")

            for info in infos:
                decoded_name = decode_answer_zip_filename(info)
                if decoded_name.endswith("/"):
                    continue
                if decoded_name in {"README.txt", ".DS_Store", "Thumbs.db"} or decoded_name.startswith("._"):
                    continue
                parts = normalize_answer_zip_path(decoded_name)
                if parts is None:
                    continue

                filename = parts[-1]
                if filename in {"README.txt", ".DS_Store", "Thumbs.db"} or filename.startswith("._"):
                    continue
                if not is_image_answer_file(filename):
                    continue

                file_count += 1
                if file_count > ANSWER_ZIP_MAX_FILES:
                    raise AnswerZipImportError("ZIP 내부 파일 수가 너무 많습니다.")
                total_bytes += info.file_size
                if total_bytes > ANSWER_ZIP_MAX_TOTAL_BYTES:
                    raise AnswerZipImportError("ZIP 내부 전체 파일 용량이 너무 큽니다.")

                chapter_number = extract_first_number(parts[0])
                item_part = parts[1] if len(parts) >= 3 else os.path.splitext(parts[-1])[0]
                item_number = extract_first_number(item_part)
                item = items_by_chapter_and_number.get((chapter_number, item_number))
                if item is None:
                    unmatched_files.append({"source_path": decoded_name, "reason": "챕터/문제 번호를 찾을 수 없습니다."})
                    continue

                matched_items.setdefault(item.id, {
                    "item_id": item.id,
                    "chapter_title": item.chapter.title,
                    "item_title": item.title,
                    "files": [],
                })["files"].append({
                    "source_path": decoded_name,
                    "zip_name": info.filename,
                    "filename": filename,
                    "size": info.file_size,
                })
    except zipfile.BadZipFile as exc:
        raise AnswerZipImportError("올바른 ZIP 파일이 아닙니다.") from exc

    for item_data in matched_items.values():
        item_data["files"].sort(key=lambda file_data: natural_sort_key(file_data["filename"]))

    if not matched_items and not unmatched_files:
        raise AnswerZipImportError("예시답안 이미지 파일을 찾지 못했습니다.")

    return {
        "import_rule": ANSWER_ZIP_OLYMPIAD_RULE,
        "items": sorted(matched_items.values(), key=lambda item_data: natural_sort_key(f"{item_data['chapter_title']} {item_data['item_title']}")),
        "unmatched_files": unmatched_files,
        "chapter_count": len({item_data["chapter_title"] for item_data in matched_items.values()}),
        "item_count": len(matched_items),
        "file_count": sum(len(item_data["files"]) for item_data in matched_items.values()),
    }


def apply_olympiad_answer_examples_zip(batch):
    preview_data = batch.preview_data or {}
    if not batch.zip_file:
        raise AnswerZipImportError("적용할 answer.zip 파일이 없습니다.")

    source_paths_by_item = {
        item_data["item_id"]: [(file_data["source_path"], file_data.get("zip_name") or file_data["source_path"]) for file_data in item_data.get("files", [])]
        for item_data in preview_data.get("items", [])
    }

    created_examples = 0
    replaced_examples = 0
    with transaction.atomic():  # type: ignore[reportGeneralTypeIssues]
        with batch.zip_file.open("rb") as zip_file:
            with zipfile.ZipFile(zip_file) as archive:
                for item_id, source_paths in source_paths_by_item.items():
                    item = Item.objects.get(id=item_id, chapter__program=batch.program, item_type="olympiad")
                    replaced_examples += item.olympiad_examples.count()
                    item.olympiad_examples.all().delete()
                    for order, (source_path, zip_name) in enumerate(source_paths, start=1):
                        content = archive.read(zip_name)
                        filename = os.path.basename(source_path)
                        example = OlympiadAnswerExample(item=item, caption=f"{item.title} 예시답안 {order}", order=order)
                        example.image.save(filename, ContentFile(content), save=True)
                        created_examples += 1

    return {
        "created_examples": created_examples,
        "replaced_examples": replaced_examples,
        "matched_items": len(source_paths_by_item),
    }


def parse_answer_zip(uploaded_file):
    uploaded_file.seek(0)
    chapters_by_title = {}
    total_bytes = 0
    total_text_bytes = 0
    file_count = 0

    try:
        with zipfile.ZipFile(uploaded_file) as archive:
            infos = archive.infolist()
            if not infos:
                raise AnswerZipImportError("비어 있는 ZIP 파일입니다.")

            for info in infos:
                decoded_name = decode_answer_zip_filename(info)
                parts = normalize_answer_zip_path(decoded_name)
                if parts is None:
                    continue

                relative_path = "/".join(parts[1:])
                is_text_file = is_text_like_answer_file(relative_path)
                is_pptx_file = is_pptx_answer_file(relative_path)

                if is_text_file and info.file_size > ANSWER_ZIP_MAX_TEXT_FILE_BYTES:
                    raise AnswerZipImportError(f"텍스트 파일이 너무 큽니다: {info.filename}")
                if is_pptx_file and info.file_size > ANSWER_ZIP_MAX_PPTX_FILE_BYTES:
                    raise AnswerZipImportError(f"PPTX 파일이 너무 큽니다: {info.filename}")
                if not is_text_file and not is_pptx_file and info.file_size > ANSWER_ZIP_MAX_TEXT_FILE_BYTES:
                    raise AnswerZipImportError(f"지원하지 않는 큰 바이너리 파일입니다: {info.filename}")

                total_bytes += info.file_size
                if total_bytes > ANSWER_ZIP_MAX_TOTAL_BYTES:
                    raise AnswerZipImportError("ZIP 내부 전체 파일 용량이 너무 큽니다.")
                file_count += 1
                if file_count > ANSWER_ZIP_MAX_FILES:
                    raise AnswerZipImportError("ZIP 내부 파일 수가 너무 많습니다.")

                is_text_file = is_text_like_answer_file(relative_path)
                answer_code = ""
                file_kind = "binary"

                if is_text_file:
                    raw_content = archive.read(info)
                    try:
                        answer_code = raw_content.decode("utf-8-sig").strip()
                        file_kind = "text"
                        total_text_bytes += info.file_size
                        if total_text_bytes > ANSWER_ZIP_MAX_TOTAL_TEXT_BYTES:
                            raise AnswerZipImportError("ZIP 내부 텍스트 총량이 너무 큽니다.")
                    except UnicodeDecodeError:
                        file_kind = "binary"

                chapter_title = parts[0].strip()
                chapter = chapters_by_title.setdefault(chapter_title, {
                    "title": chapter_title,
                    "items": [],
                })
                chapter["items"].append({
                    "title": title_from_zip_path(relative_path),
                    "key": make_answer_zip_item_key(chapter_title, relative_path),
                    "source_path": decoded_name,
                    "relative_path": relative_path,
                    "answer_code": answer_code,
                    "file_kind": file_kind,
                    "size": info.file_size,
                })
    except zipfile.BadZipFile as exc:
        raise AnswerZipImportError("올바른 ZIP 파일이 아닙니다.") from exc

    chapters = sorted(chapters_by_title.values(), key=lambda chapter: natural_sort_key(chapter["title"]))
    for chapter_index, chapter in enumerate(chapters, start=1):
        chapter["number"] = chapter_index
        chapter["items"].sort(key=lambda item: natural_sort_key(item["relative_path"]))
        for item_index, item in enumerate(chapter["items"], start=1):
            item["number"] = item_index

    if not chapters:
        raise AnswerZipImportError("A-rule에 맞는 챕터 폴더와 텍스트 파일을 찾지 못했습니다.")

    return {
        "chapters": chapters,
        "chapter_count": len(chapters),
        "item_count": sum(len(chapter["items"]) for chapter in chapters),
    }


def apply_answer_zip_preview(program, preview_data):
    with transaction.atomic():  # type: ignore[reportGeneralTypeIssues]
        created_chapters = 0
        updated_items = 0
        created_items = 0

        existing_chapters = {chapter.title: chapter for chapter in program.chapters.all()}
        next_chapter_number = (program.chapters.order_by("-number").values_list("number", flat=True).first() or 0) + 1

        for chapter_data in preview_data.get("chapters", []):
            chapter_title = chapter_data["title"]
            chapter = existing_chapters.get(chapter_title)
            if chapter is None:
                chapter = Chapter.objects.create(
                    program=program,
                    number=next_chapter_number,
                    title=chapter_title,
                    content="answer.zip에서 생성된 챕터입니다.",
                )
                existing_chapters[chapter_title] = chapter
                next_chapter_number += 1
                created_chapters += 1

            next_item_number = (chapter.items.order_by("-number").values_list("number", flat=True).first() or 0) + 1
            existing_items = {item.key: item for item in chapter.items.filter(key__startswith=f"{ANSWER_ZIP_ITEM_KEY_PREFIX}_")}

            for item_data in chapter_data.get("items", []):
                item = existing_items.get(item_data["key"])
                item_defaults = {
                    "title": item_data["title"],
                    "item_type": "problem" if item_data.get("file_kind") == "text" else "project",
                    "explain_html": (
                        f"<p>answer.zip 원본 경로: <code>{item_data['source_path']}</code></p>"
                        f"<p>파일 유형: {'텍스트' if item_data.get('file_kind') == 'text' else '바이너리/PPTX 등'}</p>"
                    ),
                    "hint": (
                        "answer.zip에서 가져온 텍스트 정답 자료입니다."
                        if item_data.get("file_kind") == "text"
                        else "answer.zip에서 가져온 바이너리 정답 자료입니다. 원본 파일 경로를 확인해 주세요."
                    ),
                    "answer_code": item_data["answer_code"],
                    "example_input": "",
                    "expected_output": "",
                }
                if item is None:
                    Item.objects.create(
                        chapter=chapter,
                        number=next_item_number,
                        key=item_data["key"],
                        **item_defaults,
                    )
                    next_item_number += 1
                    created_items += 1
                else:
                    changed_fields = []
                    for field, value in item_defaults.items():
                        if getattr(item, field) != value:
                            setattr(item, field, value)
                            changed_fields.append(field)
                    if changed_fields:
                        item.save(update_fields=changed_fields)
                        updated_items += 1

        return {
            "created_chapters": created_chapters,
            "created_items": created_items,
            "updated_items": updated_items,
        }

def is_ppt_exam_item(item):
    key = (item.key or "").lower()
    title = item.title or ""
    chapter_title = item.chapter.title or ""
    return key.startswith("itq_ppt_expected_") or key.startswith("itq_ppt_past_") or "모의고사" in title or "기출문제" in title or "모의고사" in chapter_title or "기출문제" in chapter_title


def get_ppt_exam_session_key(item_id):
    return f"ppt_exam_start_{item_id}"


def get_ppt_exam_deadline(request, item_id):
    started_at_iso = request.session.get(get_ppt_exam_session_key(item_id))
    if not started_at_iso:
        return None

    try:
        started_at = timezone.datetime.fromisoformat(started_at_iso)
    except ValueError:
        request.session.pop(get_ppt_exam_session_key(item_id), None)
        return None

    if timezone.is_naive(started_at):
        started_at = timezone.make_aware(started_at, timezone.get_current_timezone())
    return started_at + timedelta(seconds=PPT_EXAM_DURATION_SECONDS)


def build_ppt_exam_feedback(score, slide_count, text_slide_count, text_length, elapsed_seconds, rubric):
    return {
        "score": score,
        "slide_count": slide_count,
        "text_slide_count": text_slide_count,
        "text_length": text_length,
        "elapsed_seconds": elapsed_seconds,
        "target_slide_count": rubric["target_slide_count"],
        "minimum_slide_count": rubric["minimum_slide_count"],
        "minimum_text_length": rubric["minimum_text_length"],
        "minimum_text_slides": rubric["minimum_text_slides"],
    }


def get_default_ppt_exam_rubric(item):
    if (item.key or "").startswith("itq_ppt_past_"):
        return {
            "target_slide_count": 10,
            "minimum_slide_count": 8,
            "minimum_text_length": 180,
            "minimum_text_slides": 6,
        }
    return {
        "target_slide_count": 12,
        "minimum_slide_count": 10,
        "minimum_text_length": 220,
        "minimum_text_slides": 8,
    }


def get_ppt_exam_rubric(item):
    raw = (item.expected_output or "").strip()
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                default = get_default_ppt_exam_rubric(item)
                return {
                    "target_slide_count": int(parsed.get("target_slide_count", default["target_slide_count"])),
                    "minimum_slide_count": int(parsed.get("minimum_slide_count", default["minimum_slide_count"])),
                    "minimum_text_length": int(parsed.get("minimum_text_length", default["minimum_text_length"])),
                    "minimum_text_slides": int(parsed.get("minimum_text_slides", default["minimum_text_slides"])),
                }
        except (TypeError, ValueError, json.JSONDecodeError):
            pass
    return get_default_ppt_exam_rubric(item)


def extract_pptx_metrics(file_path):
    with zipfile.ZipFile(file_path) as pptx_zip:
        names = pptx_zip.namelist()
        def slide_sort_key(path):
            match = re.search(r"slide(\d+)\.xml$", path)
            return int(match.group(1)) if match else 0

        slide_paths = sorted(
            [name for name in names if name.startswith("ppt/slides/slide") and name.endswith(".xml")],
            key=slide_sort_key,
        )
        if not slide_paths:
            raise ValueError("슬라이드가 없는 PPTX 파일입니다.")

        slide_texts = []
        for slide_path in slide_paths:
            xml_bytes = pptx_zip.read(slide_path)
            root = ET.fromstring(xml_bytes)
            texts = [node.text.strip() for node in root.findall('.//a:t', PPT_SLIDE_XML_NS) if node.text and node.text.strip()]
            slide_texts.append(" ".join(texts).strip())

    joined_text = "\n".join(text for text in slide_texts if text)
    return {
        "slide_count": len(slide_paths),
        "text_slide_count": sum(1 for text in slide_texts if text),
        "text_length": len(joined_text),
    }


def grade_ppt_exam_submission(item, file_path, elapsed_seconds):
    rubric = get_ppt_exam_rubric(item)
    metrics = extract_pptx_metrics(file_path)
    score = 30

    slide_count = metrics["slide_count"]
    text_slide_count = metrics["text_slide_count"]
    text_length = metrics["text_length"]

    if slide_count >= rubric["minimum_slide_count"]:
        score += 25
    else:
        score += round(25 * max(slide_count, 0) / max(rubric["minimum_slide_count"], 1))

    if text_slide_count >= rubric["minimum_text_slides"]:
        score += 20
    else:
        score += round(20 * max(text_slide_count, 0) / max(rubric["minimum_text_slides"], 1))

    if text_length >= rubric["minimum_text_length"]:
        score += 15
    else:
        score += round(15 * max(text_length, 0) / max(rubric["minimum_text_length"], 1))

    distance = abs(slide_count - rubric["target_slide_count"])
    score += max(0, 10 - min(distance, 10))

    completed = score >= 70
    feedback = build_ppt_exam_feedback(score, slide_count, text_slide_count, text_length, elapsed_seconds, rubric)
    return score, completed, feedback

# --- 권한 체크 유틸리티 ---
def is_admin(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


@login_required
@user_passes_test(is_admin)
def download_answer_zip_template(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)
    buffer = build_olympiad_answer_template_zip(program)
    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="answer_template_program_{program.id}.zip"'
    return response


def require_full_member(view_func):
    """메듀랩 정회원만 접근 가능하도록 하는 데코레이터"""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        profile = getattr(request.user, 'profile', None)
        # 관리자는 항상 통과
        if request.user.is_staff or request.user.is_superuser:
            return view_func(request, *args, **kwargs)
        if not profile or not profile.is_full_member:
            messages.warning(request, '메듀랩 정회원만 이용 가능합니다.')
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return _wrapped

# --- Python 오류 한글 가이드 매핑 ---
PYTHON_ERROR_GUIDE = {
    "NameError": "정의되지 않은 이름을 사용했습니다. 오타가 있거나, 변수를 만들기 전에 사용한 건 아닌지 확인해 보세요.",
    "SyntaxError": "파이썬 문법에 어긋나는 부분이 있습니다. 괄호를 안 닫았거나, 콜론(:)이 빠졌는지 확인해 보세요.",
    "IndentationError": "들여쓰기(강제 공백)가 잘못되었습니다. 줄 앞의 빈칸이 일정한지 확인해 보세요.",
    "TypeError": "데이터 타입이 맞지 않습니다. 숫자와 글자를 더하려고 한 것은 아닌지 확인해 보세요.",
    "ZeroDivisionError": "0으로 나눌 수 없습니다. 나누는 값(분모)이 0이 되지 않도록 코드를 수정해 보세요.",
    "IndexError": "목록(List)의 범위를 벗어난 위치에 접근하려고 했습니다.",
    "KeyError": "딕셔너리에 존재하지 않는 키를 사용했습니다.",
    "ValueError": "부적절한 값을 사용했습니다. 글자를 숫자로 바꾸려고 한 건 아닌지 확인해 보세요.",
    "AttributeError": "해당 객체에 없는 기능을 사용하려고 했습니다.",
    "ModuleNotFoundError": "설치되지 않았거나 존재하지 않는 모듈을 불러오려고 했습니다.",
}


def get_korean_error_hint(error):
    error_type = type(error).__name__
    message = str(error)

    if error_type == "ValueError":
        lowered = message.lower()
        if "not enough values to unpack" in lowered:
            return "입력값이 부족합니다. 문제에서 요구한 개수만큼 입력했는지 확인해 보세요. 예를 들어 두 수를 입력해야 하면 `3 4`처럼 넣어야 합니다."
        if "too many values to unpack" in lowered:
            return "입력값이 너무 많습니다. 문제에서 요구한 개수보다 더 많이 입력하지 않았는지 확인해 보세요."
        if "invalid literal for int()" in lowered:
            return "숫자로 바꿀 수 없는 값을 입력했습니다. `int()`로 바꾸는 값이 숫자인지 확인해 보세요."
        if "invalid literal for float()" in lowered:
            return "실수로 바꿀 수 없는 값을 입력했습니다. `float()`로 바꾸는 값이 숫자인지 확인해 보세요."

    return PYTHON_ERROR_GUIDE.get(error_type, "오류가 발생했습니다. 코드를 다시 차근차근 확인해 보세요.")

# --- Python 코드 안전 실행 유틸리티 ---
def safe_exec(code, input_str=""):
    old_stdout = sys.stdout
    sys.stdout = captured = io.StringIO()
    
    # input() 가로채기 위한 클래스
    class MockInput:
        def __init__(self, data):
            # splitlines()를 사용하여 \r\n, \n 등 모든 종류의 줄바꿈을 깔끔하게 제거
            self.lines = data.splitlines() if data else []
            self.idx = 0
        def __call__(self, prompt=""):
            if self.idx < len(self.lines):
                res = self.lines[self.idx]
                self.idx += 1
                return res
            return ""

    try:
        # 빌트인 함수 제한 및 input 가로채기
        allowed_builtins = {
            "print": print,
            "range": range,
            "len": len,
            "int": int,
            "str": str,
            "float": float,
            "list": list,
            "tuple": tuple,
            "dict": dict,
            "set": set,
            "sum": sum,
            "abs": abs,
            "round": round,
            "map": map,
            "max": max,
            "min": min,
            "sorted": sorted,
            "enumerate": enumerate,
            "zip": zip,
            "input": MockInput(input_str),
            "type": type,
        }
        exec(code, {"__builtins__": allowed_builtins})
        output = captured.getvalue()
    except Exception as e:
        korean_hint = get_korean_error_hint(e)
        # SyntaxError/IndentationError 등은 e.msg에 메시지가 있고 str(e)에 위치가 포함됨
        output = f"Traceback (Error Notification):\n{e}\n\n[💡 도움말]\n{korean_hint}"
    finally:
        sys.stdout = old_stdout
    return output


def parse_objective_options(raw_text):
    options = []
    if not raw_text:
        return options

    for line in raw_text.replace("\r\n", "\n").split("\n"):
        text = line.strip()
        if not text:
            continue

        match = re.match(r"^([A-Za-z0-9]+)[\.)\s:-]+(.*)$", text)
        if match:
            label = match.group(1).strip().upper()
            content = match.group(2).strip()
        else:
            label = str(len(options) + 1)
            content = text

        options.append({"label": label, "text": content})

    return options


LEARNING_SECTION_LABELS = {
    "programming implementation": "문제 설명",
    "input description": "입력 형식",
    "output description": "출력 형식",
    "scoring criteria": "채점 포인트",
    "guide note": "원문 예시",
    "example note": "예시 설명",
    "rule note": "규칙 정리",
}


def build_learning_sections(explain_html):
    if not explain_html:
        return []

    sections = []
    paragraph_texts = re.findall(r"<p>(.*?)</p>", explain_html, flags=re.IGNORECASE | re.DOTALL)

    for raw_text in paragraph_texts:
        text = re.sub(r"<[^>]+>", "", raw_text)
        text = unescape(text).strip()
        if not text:
            continue

        matched = False
        for prefix, label in LEARNING_SECTION_LABELS.items():
            marker = f"{prefix}:"
            if text.lower().startswith(marker):
                body = text[len(marker):].strip()
                sections.append({"label": label, "body": body})
                matched = True
                break

        if not matched:
            sections.append({"label": "학습 내용", "body": text})

    return sections


def get_member_filter_name(user):
    profile = getattr(user, 'profile', None)
    if profile:
        return (profile.nickname or profile.real_name or user.get_full_name() or user.username).strip()
    return (user.get_full_name() or user.username).strip()


EXCEL_IMPORT_HEADER_ALIASES = {
    "chapter_number": {"장번호"},
    "chapter_title": {"장제목"},
    "chapter_description": {"장설명"},
    "item_number": {"항목순서"},
    "item_key": {"아이템키(ex01)", "아이템키"},
    "item_title": {"아이템제목"},
    "item_type": {"유형(example/problem)", "유형"},
    "explain_html": {"설명HTML"},
    "hint": {"힌트"},
    "answer_code": {"정답코드"},
    "example_input": {"예시입력"},
    "expected_output": {"예상출력"},
}


def normalize_excel_header(header_value):
    return str(header_value or "").strip()


def build_excel_header_map(header_row):
    normalized_headers = [normalize_excel_header(value) for value in header_row]
    header_map = {}
    for field_name, aliases in EXCEL_IMPORT_HEADER_ALIASES.items():
        for index, header in enumerate(normalized_headers):
            if header in aliases:
                header_map[field_name] = index
                break
    return header_map


def get_excel_cell(row, header_map, field_name, default=None):
    index = header_map.get(field_name)
    if index is None or index >= len(row):
        return default
    value = row[index]
    return default if value is None else value


def parse_course_excel_row(row, header_map):
    chapter_number = get_excel_cell(row, header_map, "chapter_number")
    chapter_title = get_excel_cell(row, header_map, "chapter_title", "")
    chapter_description = get_excel_cell(row, header_map, "chapter_description", "")
    item_number = get_excel_cell(row, header_map, "item_number", 1)
    item_key = get_excel_cell(row, header_map, "item_key", "")
    item_title = get_excel_cell(row, header_map, "item_title", "")
    item_type = get_excel_cell(row, header_map, "item_type", "example")
    explain_html = get_excel_cell(row, header_map, "explain_html", "")
    hint = get_excel_cell(row, header_map, "hint", "")
    answer_code = get_excel_cell(row, header_map, "answer_code", "")
    example_input = get_excel_cell(row, header_map, "example_input", "")
    expected_output = get_excel_cell(row, header_map, "expected_output", "")

    if chapter_number in (None, "") or not item_key or not item_title:
        return None

    try:
        chapter_number = int(chapter_number)
    except (TypeError, ValueError):
        raise ValueError(f"장번호가 올바르지 않습니다: {chapter_number}")

    try:
        item_number = int(item_number or 1)
    except (TypeError, ValueError):
        item_number = 1

    return {
        "chapter_number": chapter_number,
        "chapter_title": str(chapter_title).strip(),
        "chapter_description": str(chapter_description or "").strip(),
        "item_number": item_number,
        "item_key": str(item_key).strip(),
        "item_title": str(item_title).strip(),
        "item_type": str(item_type or "example").strip() or "example",
        "explain_html": str(explain_html or ""),
        "hint": str(hint or ""),
        "answer_code": str(answer_code or ""),
        "example_input": str(example_input or ""),
        "expected_output": str(expected_output or ""),
    }

# --- (1) 관리자용 학습 프로그램 목록 ---
@login_required
@user_passes_test(is_admin)
def learning_program_list(request):
    programs = LearningProgram.objects.all().order_by("-created_at")
    return render(request, "courses/learning_program_list.html", {
        "programs": programs
    })

# --- (1-1) 과정 등록 ---
@login_required
@user_passes_test(is_admin)
def learning_program_create(request):
    if request.method == "POST":
        form = CourseForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "새 과정이 성공적으로 등록되었습니다.")
            return redirect("student_course_list")
    else:
        form = CourseForm()
    
    return render(request, "courses/learning_program_form.html", {
        "form": form,
        "title": "새 과정 등록"
    })

# --- (1-2) 과정 수정 ---
@login_required
@user_passes_test(is_admin)
def learning_program_edit(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)
    if request.method == "POST":
        form = CourseForm(request.POST, request.FILES, instance=program)
        if form.is_valid():
            form.save()
            messages.success(request, f"'{program.name}' 과정이 수정되었습니다.")
            return redirect("student_course_list")
    else:
        form = CourseForm(instance=program)
    
    return render(request, "courses/learning_program_form.html", {
        "form": form,
        "title": "과정 정보 수정",
        "program": program
    })

# --- (1-3) 과정 삭제 ---
@login_required
@user_passes_test(is_admin)
def learning_program_delete(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)
    if request.method == "POST":
        name = program.name
        program.delete()
        messages.success(request, f"'{name}' 과정이 삭제되었습니다.")
        return redirect("student_course_list")
    return render(request, "courses/learning_program_confirm_delete.html", {"program": program})

# --- (1-4) 유형 관리 목록 ---
@login_required
@user_passes_test(is_admin)
def program_type_list(request):
    types = ProgramType.objects.all().order_by("order", "name")
    return render(request, "courses/program_type_list.html", {"types": types})

# --- (1-5) 유형 등록 ---
@login_required
@user_passes_test(is_admin)
def program_type_create(request):
    if request.method == "POST":
        form = ProgramTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "새 과정 유형이 등록되었습니다.")
            return redirect("program_type_list")
    else:
        form = ProgramTypeForm()
    return render(request, "courses/program_type_form.html", {"form": form, "title": "새 유형 등록"})

# --- (1-6) 유형 수정 ---
@login_required
@user_passes_test(is_admin)
def program_type_edit(request, type_id):
    p_type = get_object_or_404(ProgramType, id=type_id)
    if request.method == "POST":
        form = ProgramTypeForm(request.POST, instance=p_type)
        if form.is_valid():
            form.save()
            messages.success(request, f"'{p_type.name}' 유형이 수정되었습니다.")
            return redirect("program_type_list")
    else:
        form = ProgramTypeForm(instance=p_type)
    return render(request, "courses/program_type_form.html", {"form": form, "title": "유형 수정"})

# --- (1-7) 유형 삭제 ---
@login_required
@user_passes_test(is_admin)
def program_type_delete(request, type_id):
    p_type = get_object_or_404(ProgramType, id=type_id)
    if request.method == "POST":
        p_type.delete()
        messages.success(request, "유형이 삭제되었습니다.")
        return redirect("program_type_list")
    return render(request, "courses/program_type_confirm_delete.html", {"type": p_type})

# --- (1-8) 엑셀 템플릿 다운로드 ---
@login_required
@user_passes_test(is_admin)
def download_course_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CourseTemplate"
    
    headers = [
        '장번호', '장제목', '장설명',
        '항목순서', '아이템키', '아이템제목', '유형',
        '설명HTML', '힌트', '정답코드', '예시입력', '예상출력'
    ]
    ws.append(headers)
    
    # 샘플 데이터 한 줄
    ws.append([1, '파이썬 기초', '기본 문법을 배웁니다', 1, 'ex01', 'Hello World', 'example', '<p>화면에 출력해보세요</p>', 'print 사용', 'print("Hello")', '', 'Hello'])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=medulab_course_template.xlsx'
    wb.save(response)
    return response

# --- (1-9) 특정 과정 데이터 엑셀로 내보내기 ---
@login_required
@user_passes_test(is_admin)
def export_program_to_excel(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CourseData"

    headers = [
        '장번호', '장제목', '장설명',
        '항목순서', '아이템키', '아이템제목', '유형',
        '설명HTML', '힌트', '정답코드', '예시입력', '예상출력'
    ]
    ws.append(headers)

    for chapter in program.chapters.all():
        for item in chapter.items.all():
            ws.append([
                chapter.number,
                chapter.title,
                chapter.content or '',
                item.number,
                item.key,
                item.title,
                item.item_type,
                item.explain_html or '',
                item.hint or '',
                item.answer_code or '',
                item.example_input or '',
                item.expected_output or '',
            ])

    safe_name = program.name.replace(' ', '_').replace('/', '_')
    filename = f"medulab_course_{safe_name}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# --- (2) 학생용 나의 코스 목록 ---
def student_course_list(request):
    selected_program_type = request.GET.get("program_type", "").strip()

    # 모든 활성 프로그램
    all_programs = LearningProgram.objects.filter(is_active=True).select_related("program_type")

    available_program_types = ProgramType.objects.filter(
        learning_programs__is_active=True
    ).distinct().order_by("order", "name")

    coding_type = available_program_types.filter(name="코딩").first()
    python_coding_filter = coding_type is not None

    # "대회 대비"와 "대회 준비"를 동일 필터로 묶기 위해 대비 타입 ID 수집
    contest_prep_ids = list(
        ProgramType.objects.filter(name__in=["대회 대비", "대회 준비"]).values_list("id", flat=True)
    )

    if selected_program_type == "python-coding":
        all_programs = all_programs.filter(
            Q(program_type=coding_type)
            | (
                (Q(program_type__name__icontains="대회") | Q(program_type__name__icontains="자격증"))
                & (Q(name__icontains="python") | Q(name__icontains="파이썬") | Q(description__icontains="python") | Q(description__icontains="파이썬"))
            )
        )
    elif selected_program_type.isdigit() and int(selected_program_type) in contest_prep_ids:
        all_programs = all_programs.filter(program_type_id__in=contest_prep_ids)
    elif selected_program_type.isdigit():
        all_programs = all_programs.filter(program_type_id=int(selected_program_type))

    all_programs = all_programs.order_by("id")
    
    # 내가 수강 중인 프로그램 ID 목록
    if request.user.is_authenticated:
        enrolled_ids = LearningEnrollment.objects.filter(user=request.user)\
                                               .values_list("program_id", flat=True)
        profile = getattr(request.user, 'profile', None)
        can_apply = bool(profile and profile.is_full_member)
    else:
        enrolled_ids = []
        can_apply = False
    
    return render(request, "courses/student_course_list.html", {
        "programs": all_programs,
        "enrolled_ids": enrolled_ids,
        "program_types": available_program_types,
        "selected_program_type": selected_program_type,
        "show_python_coding_filter": python_coding_filter,
        "can_apply": can_apply,
    })

# --- (3) 수강 신청 ---
@login_required
@require_full_member
def student_course_apply(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)
    
    # 중복 체크
    if LearningEnrollment.objects.filter(user=request.user, program=program).exists():
        messages.warning(request, "이미 신청한 과정입니다.")
    else:
        LearningEnrollment.objects.create(user=request.user, program=program)
        messages.success(request, f"'{program.name}' 수강 신청이 완료되었습니다!")
        
    return redirect("course_home", program_id=program.id)

# --- (4) 코스 홈 (챕터 구성) ---
def course_home(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)
    chapters = Chapter.objects.filter(program=program).order_by("number")
    program_badge = get_program_completion_badge(program)
    if request.user.is_authenticated:
        program_badge_awarded = request.user.earned_badges.filter(badge=program_badge).exists()
        recent_badges = get_recent_user_badges(request.user, limit=4)
    else:
        program_badge_awarded = False
        recent_badges = []
    
    # 챕터별 진도율 계산하여 챕터 객체에 주입
    for ch in chapters:
        total_items = Item.objects.filter(chapter=ch).count()
        if request.user.is_authenticated:
            completed_items = UserProgress.objects.filter(
                user=request.user, 
                item__chapter=ch, 
                completed=True
            ).count()
        else:
            completed_items = 0
        ch.progress = round((completed_items / total_items * 100)) if total_items > 0 else 0
        
    return render(request, "learning_program/course_home.html", {
        "program": program,
        "chapters": chapters,
        "program_badge": program_badge,
        "program_badge_awarded": program_badge_awarded,
        "recent_badges": recent_badges,
    })

# --- (5) 챕터 상세 (항목 목록) ---
def chapter_detail(request, chapter_id):
    chapter = get_object_or_404(Chapter, id=chapter_id)
    items = list(Item.objects.filter(chapter=chapter).order_by("number"))

    if request.user.is_authenticated:
        progress_map = {p.item_id: p.completed for p in UserProgress.objects.filter(user=request.user, item__in=items)}
    else:
        progress_map = {}
    for item in items:
        item.is_completed = progress_map.get(item.id, False)

    # olympiad 아이템은 하위문제별 카드로 확장
    sub_answer_map = {}
    olympiad_ids = [it.id for it in items if it.item_type == "olympiad"]
    if olympiad_ids:
        sqs_all = list(OlympiadSubQuestion.objects.filter(item_id__in=olympiad_ids).order_by("item__number", "number"))
        if request.user.is_authenticated and sqs_all:
            sq_ids = [sq.id for sq in sqs_all]
            sub_answer_map = {
                sa.sub_question_id: sa
                for sa in OlympiadSubAnswer.objects.filter(sub_question_id__in=sq_ids, student=request.user)
            }

    # 카드 목록 구성 (olympiad → 하위문제 카드, 나머지 → 원래 아이템)
    card_list = []
    for item in items:
        if item.item_type == "olympiad":
            sqs = [sq for sq in sqs_all if sq.item_id == item.id] if olympiad_ids else []
            if sqs:
                for sq in sqs:
                    my_ans = sub_answer_map.get(sq.id)
                    card_list.append({
                        "type": "sub_question",
                        "item": item,
                        "sq": sq,
                        "is_completed": bool(my_ans and (my_ans.text_answer or my_ans.photo)),
                        "item_id": item.id,
                        "sq_number": sq.number,
                    })
            else:
                card_list.append({"type": "item", "item": item, "is_completed": item.is_completed})
        else:
            card_list.append({"type": "item", "item": item, "is_completed": item.is_completed})

    return render(request, "learning_program/chapter_detail.html", {
        "chapter": chapter,
        "items": items,
        "card_list": card_list,
        "program": chapter.program
    })

# --- (6) 학습 아이템 페이지 ---
def item_page(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    program = item.chapter.program
    
    # 이전/다음 이동 로직 (챕터 내 또는 전체 과정 내)
    all_items = list(Item.objects.filter(chapter__program=program).order_by('chapter__number', 'number'))
    try:
        current_idx = all_items.index(item)
        prev_item = all_items[current_idx - 1] if current_idx > 0 else None
        next_item = all_items[current_idx + 1] if current_idx < len(all_items) - 1 else None
    except ValueError:
        prev_item = next_item = None

    # 유저 진행 상황 (기존 코드 등 로드)
    if request.user.is_authenticated:
        progress, _ = UserProgress.objects.get_or_create(user=request.user, item=item)
    else:
        progress = SimpleNamespace(code="", last_output="", score=0, completed=False)
    objective_options = parse_objective_options(item.example_input) if item.item_type == "objective" else []
    learning_sections = build_learning_sections(item.explain_html)
    is_olympiad = item.item_type == "olympiad"
    olympiad_submission = None
    olympiad_submission_form = None
    olympiad_examples = []
    olympiad_feedback_sections = []

    if is_olympiad and request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        olympiad_submission = OlympiadAnswerSubmission.objects.filter(item=item, student=request.user).first()
        olympiad_submission_form = OlympiadAnswerSubmissionForm(instance=olympiad_submission)
        if olympiad_submission and olympiad_submission.feedback:
            olympiad_feedback_sections = parse_olympiad_feedback(olympiad_submission.feedback)

    if is_olympiad and (olympiad_submission or (request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser))):
        olympiad_examples = item.olympiad_examples.all()

    # 하위 문제 및 학생 답안 로드 (dict 리스트로 전달 — 언더스코어 속성 불가)
    sub_questions = []
    if is_olympiad:
        sqs = list(item.sub_questions.all())
        my_answers = {}
        if sqs and request.user.is_authenticated:
            my_answers = {
                sa.sub_question_id: sa
                for sa in OlympiadSubAnswer.objects.filter(
                    sub_question__item=item, student=request.user
                )
            }
        sub_questions = [
            {"sq": sq, "my_answer": my_answers.get(sq.id)}
            for sq in sqs
        ]

    # 템플릿 결정 (과정 이름이나 유형에 따라 분기 가능)
    template_name = "learning_program/item_page.html"
    p_name = program.name.lower()
    p_type_name = program.program_type.name.lower() if program.program_type else ""
    is_ppt_exam = is_ppt_exam_item(item)
    
    if is_olympiad:
        template_name = "learning_program/item_page_olympiad.html"
    elif is_ppt_exam:
        template_name = "learning_program/item_page_ppt.html"
    elif "python" in p_name or "파이썬" in p_name or "python" in p_type_name or "파이썬" in p_type_name:
        template_name = "learning_program/item_page_python.html"
    elif item.answer_code or item.expected_output:
        # 데이터가 있으면 실습 문항으로 간주하여 파이썬 페이지 노출
        template_name = "learning_program/item_page_python.html"
    elif "ppt" in p_name or "파워포인트" in p_name or "ppt" in p_type_name or "파워포인트" in p_type_name:
        template_name = "learning_program/item_page_ppt.html"

    deadline_at = get_ppt_exam_deadline(request, item.id) if is_ppt_exam else None
    now = timezone.now() if is_ppt_exam else None
    remaining_seconds = 0
    exam_active = False
    feedback_data = None
    if deadline_at and now:
        remaining_seconds = max(0, int((deadline_at - now).total_seconds()))
        exam_active = remaining_seconds > 0
    if is_ppt_exam and getattr(progress, 'last_output', ''):
        try:
            feedback_data = json.loads(progress.last_output)
        except (TypeError, ValueError, json.JSONDecodeError):
            feedback_data = None

    return render(request, template_name, {
        "item": item,
        "prev_item": prev_item,
        "next_item": next_item,
        "program": program,
        "user_progress": progress,
        "objective_options": objective_options,
        "learning_sections": learning_sections,
        "objective_selected": progress.code.strip(),
        "is_objective": item.item_type == "objective",
        "is_ppt_exam": is_ppt_exam,
        "ppt_exam_duration_seconds": PPT_EXAM_DURATION_SECONDS,
        "ppt_exam_deadline_iso": deadline_at.isoformat() if deadline_at and exam_active else "",
        "ppt_exam_remaining_seconds": remaining_seconds,
        "ppt_exam_active": exam_active,
        "ppt_exam_score": progress.score,
        "ppt_exam_completed": progress.completed,
        "ppt_exam_feedback": feedback_data,
        "is_olympiad": is_olympiad,
        "olympiad_submission": olympiad_submission,
        "olympiad_submission_form": olympiad_submission_form,
        "olympiad_examples": olympiad_examples,
        "olympiad_feedback_sections": olympiad_feedback_sections,
        "sub_questions": sub_questions,
    })


@login_required
@require_full_member
def submit_olympiad_answer(request, item_id):
    item = get_object_or_404(Item, id=item_id, item_type="olympiad")
    program = item.chapter.program

    if request.method != "POST":
        return redirect("item_page", item_id=item.id)

    if request.user.is_staff or request.user.is_superuser:
        messages.warning(request, "관리자는 올림피아드 답안을 직접 제출할 수 없습니다.")
        return redirect("item_page", item_id=item.id)

    if not LearningEnrollment.objects.filter(user=request.user, program=program).exists():
        messages.warning(request, "수강 중인 과정의 올림피아드 답안만 제출할 수 있습니다.")
        return redirect("course_home", program_id=program.id)

    submission = OlympiadAnswerSubmission.objects.filter(item=item, student=request.user).first()
    form = OlympiadAnswerSubmissionForm(request.POST, request.FILES, instance=submission)
    if form.is_valid():
        olympiad_submission = form.save(commit=False)
        olympiad_submission.item = item
        olympiad_submission.student = request.user
        olympiad_submission.status = OlympiadAnswerSubmission.STATUS_SUBMITTED
        try:
            olympiad_submission.feedback = build_olympiad_submission_feedback(olympiad_submission, item)
        except Exception:
            olympiad_submission.feedback = (
                "AI 평가점수\n"
                "- 평가 보류\n"
                "\n"
                "답안 보완 포인트\n"
                "- 답안 사진은 정상 제출되었습니다. 자동 보완 피드백을 만들지 못했지만, 예시답안과 문제 힌트를 보며 조건, 풀이 과정, 결론이 모두 들어갔는지 확인해 주세요."
            )
        olympiad_submission.save()

        progress, _ = UserProgress.objects.get_or_create(user=request.user, item=item)
        progress.code = "OLYMPIAD_SUBMITTED"
        progress.completed = True
        progress.last_output = "올림피아드 답안이 제출되었습니다."
        progress.save(update_fields=["code", "completed", "last_output", "updated_at"])
        messages.success(request, "올림피아드 답안이 제출되었습니다.")
    else:
        messages.error(request, "답안 제출 내용을 다시 확인해 주세요.")

    return redirect("item_page", item_id=item.id)


def sub_question_page(request, item_id, sq_number):
    """하위문제 전용 페이지 — 문제 하나만 표시, 이전/다음 네비게이션"""
    item = get_object_or_404(Item, id=item_id, item_type="olympiad")
    sq = get_object_or_404(OlympiadSubQuestion, item=item, number=sq_number)

    all_sqs = list(item.sub_questions.order_by("number"))
    idx = next((i for i, s in enumerate(all_sqs) if s.id == sq.id), 0)
    prev_sq = all_sqs[idx - 1] if idx > 0 else None
    next_sq = all_sqs[idx + 1] if idx < len(all_sqs) - 1 else None

    my_answer = None
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        my_answer = OlympiadSubAnswer.objects.filter(sub_question=sq, student=request.user).first()
        if my_answer and not my_answer.upload_token:
            my_answer.upload_token = secrets.token_urlsafe(32)
            my_answer.save(update_fields=["upload_token"])

    # explain_html에서 하위문제 <p> 태그 제거 — 서문(공통 안내)만 남김
    import re as _re
    intro_html = _re.sub(
        r'<p>\s*\d+-\d+\..*?</p>', '', item.explain_html or '', flags=_re.DOTALL
    ).strip()

    return render(request, "learning_program/sub_question_page.html", {
        "item": item,
        "sq": sq,
        "prev_sq": prev_sq,
        "next_sq": next_sq,
        "all_sqs": all_sqs,
        "sq_index": idx + 1,
        "sq_total": len(all_sqs),
        "my_answer": my_answer,
        "intro_html": intro_html,
        "program": item.chapter.program,
    })


@login_required
@require_full_member
def submit_sub_answer(request, sub_question_id):
    """하위 문제 텍스트 서술형 + 사진 제출"""
    sq = get_object_or_404(OlympiadSubQuestion, id=sub_question_id)
    item = sq.item
    program = item.chapter.program

    if request.method != "POST":
        return redirect("item_page", item_id=item.id)

    if request.user.is_staff or request.user.is_superuser:
        messages.warning(request, "관리자는 직접 답안을 제출할 수 없습니다.")
        return redirect("item_page", item_id=item.id)

    if not LearningEnrollment.objects.filter(user=request.user, program=program).exists():
        messages.warning(request, "수강 중인 과정의 답안만 제출할 수 있습니다.")
        return redirect("item_page", item_id=item.id)

    answer, _ = OlympiadSubAnswer.objects.get_or_create(
        sub_question=sq, student=request.user,
        defaults={"upload_token": secrets.token_urlsafe(32)}
    )
    if not answer.upload_token:
        answer.upload_token = secrets.token_urlsafe(32)

    text = request.POST.get("text_answer", "").strip()
    photo = request.FILES.get("photo")

    answer.text_answer = text
    if photo:
        answer.photo = photo
        answer.ai_score = None  # 사진 바뀌면 재분석
        answer.ai_feedback = ""
    answer.save()

    # AI 분석 (API 키 있을 때만)
    from .ai_service import run_ai_analysis
    if getattr(settings, 'ANTHROPIC_API_KEY', ''):
        try:
            run_ai_analysis(answer)
        except Exception:
            pass  # AI 실패해도 저장은 완료

    # 모든 하위문제에 답안이 있으면 아이템 완료로 표시
    total_sq = item.sub_questions.count()
    answered = OlympiadSubAnswer.objects.filter(
        sub_question__item=item, student=request.user
    ).exclude(text_answer="", photo="").count()
    if total_sq > 0 and answered >= total_sq:
        progress, _ = UserProgress.objects.get_or_create(user=request.user, item=item)
        if not progress.completed:
            progress.code = "OLYMPIAD_SUBMITTED"
            progress.completed = True
            progress.last_output = "하위 문제 답안 제출 완료"
            progress.save(update_fields=["code", "completed", "last_output", "updated_at"])

    messages.success(request, f"{item.number}-{sq.number} 답안이 저장되었습니다.")
    return redirect("sub_question_page", item_id=item.id, sq_number=sq.number)


def olympiad_qr_upload_page(request, token):
    """QR 코드 스캔 후 접속하는 모바일 사진 업로드 페이지"""
    answer = get_object_or_404(OlympiadSubAnswer, upload_token=token)
    sq = answer.sub_question
    item = sq.item

    if request.method == "POST":
        photo = request.FILES.get("photo")
        if photo:
            answer.photo = photo
            answer.ai_score = None
            answer.ai_feedback = ""
            answer.save(update_fields=["photo", "ai_score", "ai_feedback", "updated_at"])
            from .ai_service import run_ai_analysis
            if getattr(settings, 'ANTHROPIC_API_KEY', ''):
                try:
                    run_ai_analysis(answer)
                except Exception:
                    pass
            return render(request, "courses/olympiad_qr_upload_done.html", {"sq": sq, "item": item})
        messages.error(request, "사진을 선택해 주세요.")

    return render(request, "courses/olympiad_qr_upload.html", {"answer": answer, "sq": sq, "item": item})


def olympiad_qr_code(request, sub_answer_id):
    """하위 문제 답안의 QR 코드 이미지(PNG) 반환"""
    if not request.user.is_authenticated:
        from django.contrib.auth.views import redirect_to_login
        return redirect_to_login(request.get_full_path())

    answer = get_object_or_404(OlympiadSubAnswer, id=sub_answer_id, student=request.user)
    if not answer.upload_token:
        answer.upload_token = secrets.token_urlsafe(32)
        answer.save(update_fields=["upload_token"])

    upload_url = request.build_absolute_uri(
        reverse("olympiad_qr_upload_page", kwargs={"token": answer.upload_token})
    )
    img = qrcode.make(upload_url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return HttpResponse(buf.getvalue(), content_type="image/png")


@login_required
def ensure_sub_answer_token(request, sub_question_id):
    """AJAX: 하위 문제 업로드 토큰 생성 후 QR 이미지 URL 반환"""
    sq = get_object_or_404(OlympiadSubQuestion, id=sub_question_id)
    if request.user.is_staff or request.user.is_superuser:
        return JsonResponse({"error": "관리자는 사용 불가"}, status=403)

    answer, _ = OlympiadSubAnswer.objects.get_or_create(
        sub_question=sq, student=request.user,
        defaults={"upload_token": secrets.token_urlsafe(32)}
    )
    if not answer.upload_token:
        answer.upload_token = secrets.token_urlsafe(32)
        answer.save(update_fields=["upload_token"])

    qr_url = reverse("olympiad_qr_code", kwargs={"sub_answer_id": answer.id})
    return JsonResponse({"qr_url": qr_url, "answer_id": answer.id})


@login_required
def start_ppt_exam(request, item_id):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)

    item = get_object_or_404(Item, id=item_id)
    if not is_ppt_exam_item(item):
        return JsonResponse({"error": "PPT 실전 시험 항목이 아닙니다."}, status=400)

    started_at = timezone.now()
    deadline_at = started_at + timedelta(seconds=PPT_EXAM_DURATION_SECONDS)
    request.session[get_ppt_exam_session_key(item.id)] = started_at.isoformat()
    request.session.modified = True
    return JsonResponse({
        "started": True,
        "duration_seconds": PPT_EXAM_DURATION_SECONDS,
        "deadline_at": deadline_at.isoformat(),
    })


@login_required
def submit_ppt_exam(request, item_id):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)

    item = get_object_or_404(Item, id=item_id)
    if not is_ppt_exam_item(item):
        return JsonResponse({"error": "PPT 실전 시험 항목이 아닙니다."}, status=400)

    deadline_at = get_ppt_exam_deadline(request, item.id)
    if not deadline_at:
        return JsonResponse({"error": "시험이 시작되지 않았습니다. 먼저 PowerPoint를 열고 시험 시작을 확인해 주세요."}, status=400)

    remaining_seconds = int((deadline_at - timezone.now()).total_seconds())
    if remaining_seconds <= 0:
        request.session.pop(get_ppt_exam_session_key(item.id), None)
        return JsonResponse({"error": "시험 시간이 종료되었습니다. 다시 시작 후 제출해 주세요."}, status=400)

    uploaded_file = request.FILES.get("pptx_file")
    if not uploaded_file:
        return JsonResponse({"error": "업로드할 PPTX 파일을 선택해 주세요."}, status=400)

    filename = (uploaded_file.name or "").lower()
    if not filename.endswith(".pptx"):
        return JsonResponse({"error": "PPTX 파일만 업로드할 수 있습니다."}, status=400)

    temp_path = ""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pptx") as temp_file:
            for chunk in uploaded_file.chunks():
                temp_file.write(chunk)
            temp_path = temp_file.name

        elapsed_seconds = PPT_EXAM_DURATION_SECONDS - max(remaining_seconds, 0)
        score, completed, feedback = grade_ppt_exam_submission(item, temp_path, elapsed_seconds)
        progress, _ = UserProgress.objects.get_or_create(user=request.user, item=item)
        progress.code = "PPTX_SUBMITTED"
        progress.score = score
        progress.completed = completed
        progress.last_output = json.dumps(feedback, ensure_ascii=False)
        progress.save()

        new_badges = []
        if completed:
            new_badges.extend(evaluate_mission_badges(request.user))
            new_badges.extend(evaluate_program_badges(request.user, item.chapter.program))

        request.session.pop(get_ppt_exam_session_key(item.id), None)
        return JsonResponse({
            "score": score,
            "completed": completed,
            "feedback": feedback,
            "new_badges": new_badges,
            "message": "채점이 완료되었습니다. 업로드된 파일은 자동 삭제되었습니다.",
        })
    except zipfile.BadZipFile:
        return JsonResponse({"error": "올바른 PPTX 파일이 아닙니다. 다시 확인해 주세요."}, status=400)
    except ValueError as exc:
        return JsonResponse({"error": str(exc)}, status=400)
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

# --- (7) 코드 채점 API ---
@login_required
def grade_code(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=400)
    
    item_id = request.POST.get("item_id")
    code = request.POST.get("code", "")
    input_str = request.POST.get("input", "")
    
    item = get_object_or_404(Item, id=item_id)

    if item.item_type == "objective":
        selected = (request.POST.get("selected_option") or code or "").strip().upper()
        correct = (item.answer_code or "").strip().upper()
        is_correct = selected == correct and bool(selected)
        score = 100 if is_correct else 0
        explanation = item.expected_output or ""
        output_message = explanation if is_correct else f"선택한 답: {selected or '-'}"

        progress, _ = UserProgress.objects.get_or_create(user=request.user, item=item)
        progress.code = selected
        progress.last_output = output_message
        progress.score = score
        progress.completed = is_correct
        progress.save()
        new_badges = []
        if is_correct:
            new_badges.extend(evaluate_mission_badges(request.user))
            new_badges.extend(evaluate_program_badges(request.user, item.chapter.program))

        return JsonResponse({
            "is_correct": is_correct,
            "output": output_message,
            "expected": correct,
            "score": score,
            "selected": selected,
            "new_badges": new_badges,
        })

    output = safe_exec(code, input_str)
    
    # 정답 비교 (줄바꿈 문자 정규화 및 공백 제거)
    def normalize(s):
        if not s: return ""
        # \r\n을 \n으로 통일하고 양끝 공백 제거
        return "\n".join([line.strip() for line in s.replace("\r\n", "\n").strip().split("\n")]).strip()

    user_out = normalize(output)
    expected_out = normalize(item.expected_output)
    
    is_correct = (user_out == expected_out)
    score = 100 if is_correct else 0
    
    # 결과 저장
    progress, _ = UserProgress.objects.get_or_create(user=request.user, item=item)
    progress.code = code
    progress.last_output = output
    progress.score = score
    progress.completed = is_correct
    progress.save()
    new_badges = []
    if is_correct:
        new_badges.extend(evaluate_mission_badges(request.user))
        new_badges.extend(evaluate_program_badges(request.user, item.chapter.program))
    
    return JsonResponse({
        "is_correct": is_correct,
        "output": output,
        "expected": expected_out,
        "score": score,
        "new_badges": new_badges,
    })

# --- (8) 챕터 관리 및 엑셀 업로드 ---
@login_required
@user_passes_test(is_admin)
def chapter_manage(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)
    answer_zip_form = AnswerZipImportForm()
    answer_zip_batch = AnswerZipImportBatch.objects.filter(
        program=program,
        status=AnswerZipImportBatch.STATUS_PREVIEW,
    ).first()
    answer_zip_preview = answer_zip_batch.preview_data if answer_zip_batch else None
    
    if request.method == "POST" and request.POST.get("import_action") == "preview_answer_zip":
        answer_zip_form = AnswerZipImportForm(request.POST, request.FILES)
        if answer_zip_form.is_valid():
            try:
                uploaded_answer_zip = answer_zip_form.cleaned_data["answer_zip"]
                if program.chapters.filter(items__item_type="olympiad").exists():
                    answer_zip_preview = parse_olympiad_answer_examples_zip(program, uploaded_answer_zip)
                    message = (
                        f"예시답안 이미지 {answer_zip_preview['file_count']}개, "
                        f"문제 {answer_zip_preview['item_count']}개 미리보기"
                    )
                else:
                    answer_zip_preview = parse_answer_zip(uploaded_answer_zip)
                    message = (
                        f"챕터 {answer_zip_preview['chapter_count']}개, "
                        f"아이템 {answer_zip_preview['item_count']}개 미리보기"
                    )
                uploaded_answer_zip.seek(0)
                answer_zip_batch = AnswerZipImportBatch.objects.create(
                    program=program,
                    zip_file=uploaded_answer_zip,
                    import_rule=answer_zip_preview.get("import_rule", "top_level_folder_chapter"),
                    preview_data=answer_zip_preview,
                    message=message,
                    created_by=request.user,
                )
                messages.success(
                    request,
                    f"answer.zip 미리보기 준비 완료: {message}",
                )
            except AnswerZipImportError as exc:
                answer_zip_preview = None
                messages.error(request, str(exc))
        else:
            messages.error(request, "answer.zip 파일을 다시 확인해 주세요.")
    elif request.method == "POST" and request.POST.get("import_action") == "apply_answer_zip":
        if not answer_zip_batch or not answer_zip_preview:
            messages.error(request, "적용할 answer.zip 미리보기 데이터가 없습니다. 먼저 ZIP을 업로드해 주세요.")
        else:
            if answer_zip_batch.import_rule == ANSWER_ZIP_OLYMPIAD_RULE:
                result = apply_olympiad_answer_examples_zip(answer_zip_batch)
                answer_zip_batch.message = (
                    f"예시답안 {result['created_examples']}개 생성, "
                    f"기존 예시 {result['replaced_examples']}개 교체"
                )
                success_message = f"answer.zip 적용 완료: {answer_zip_batch.message}"
            else:
                result = apply_answer_zip_preview(program, answer_zip_preview)
                answer_zip_batch.message = (
                    f"챕터 {result['created_chapters']}개 생성, "
                    f"아이템 {result['created_items']}개 생성, {result['updated_items']}개 업데이트"
                )
                success_message = (
                    "answer.zip 적용 완료: "
                    f"챕터 {result['created_chapters']}개 생성, "
                    f"아이템 {result['created_items']}개 생성, {result['updated_items']}개 업데이트"
                )
            answer_zip_batch.status = AnswerZipImportBatch.STATUS_APPLIED
            answer_zip_batch.applied_at = timezone.now()
            answer_zip_batch.save(update_fields=["status", "applied_at", "message"])
            answer_zip_preview = None
            answer_zip_batch = None
            messages.success(request, success_message)
    elif request.method == "POST" and request.POST.get("import_action") == "clear_answer_zip_preview":
        if answer_zip_batch:
            answer_zip_batch.status = AnswerZipImportBatch.STATUS_FAILED
            answer_zip_batch.message = "관리자가 미리보기를 취소했습니다."
            answer_zip_batch.save(update_fields=["status", "message"])
        answer_zip_preview = None
        answer_zip_batch = None
        messages.info(request, "answer.zip 미리보기를 취소했습니다.")
    elif request.method == "POST" and request.FILES.get("excel_file"):
        file = request.FILES["excel_file"]
        try:
            wb = openpyxl.load_workbook(file)
            # 기존 데이터 유지 혹은 삭제 (여기서는 덮어쓰기 개념으로 기존 아이템 삭제)
            Chapter.objects.filter(program=program).delete()
            
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not header_values:
                    continue

                header_map = build_excel_header_map(header_values)
                required_headers = {"chapter_number", "chapter_title", "chapter_description", "item_key", "item_title", "item_type", "explain_html", "hint", "answer_code", "expected_output"}
                missing_headers = sorted(required_headers - set(header_map.keys()))
                if missing_headers:
                    raise ValueError(f"엑셀 헤더가 올바르지 않습니다. 누락: {', '.join(missing_headers)}")

                for row in ws.iter_rows(min_row=2, values_only=True):
                    parsed_row = parse_course_excel_row(row, header_map)
                    if not parsed_row:
                        continue

                    chapter, _ = Chapter.objects.get_or_create(
                        program=program, 
                        number=parsed_row["chapter_number"],
                        defaults={"title": parsed_row["chapter_title"], "content": parsed_row["chapter_description"]}
                    )

                    chapter_updates = []
                    if chapter.title != parsed_row["chapter_title"]:
                        chapter.title = parsed_row["chapter_title"]
                        chapter_updates.append("title")
                    if (chapter.content or "") != parsed_row["chapter_description"]:
                        chapter.content = parsed_row["chapter_description"]
                        chapter_updates.append("content")
                    if chapter_updates:
                        chapter.save(update_fields=chapter_updates)
                    
                    Item.objects.create(
                        chapter=chapter,
                        number=parsed_row["item_number"],
                        key=parsed_row["item_key"],
                        title=parsed_row["item_title"],
                        item_type=parsed_row["item_type"] or 'example',
                        explain_html=parsed_row["explain_html"],
                        hint=parsed_row["hint"],
                        answer_code=parsed_row["answer_code"],
                        example_input=parsed_row["example_input"],
                        expected_output=parsed_row["expected_output"]
                    )
            messages.success(request, "엑셀 데이터를 성공적으로 불러왔습니다.")
        except Exception as e:
            messages.error(request, f"엑셀 처리 중 오류 발생: {str(e)}")
            
    chapters = Chapter.objects.filter(program=program).prefetch_related('items')
    return render(request, "courses/chapter_manage.html", {
        "program": program,
        "chapters": chapters,
        "answer_zip_form": answer_zip_form,
        "answer_zip_batch": answer_zip_batch,
        "answer_zip_preview": answer_zip_preview,
    })


@login_required
@user_passes_test(is_admin)
def answer_zip_apply(request, program_id, batch_id):
    program = get_object_or_404(LearningProgram, id=program_id)
    batch = get_object_or_404(
        AnswerZipImportBatch,
        id=batch_id,
        program=program,
        status=AnswerZipImportBatch.STATUS_PREVIEW,
    )
    if request.method != "POST":
        return redirect("chapter_manage", program_id=program.id)

    if batch.import_rule == ANSWER_ZIP_OLYMPIAD_RULE:
        result = apply_olympiad_answer_examples_zip(batch)
        batch.message = (
            f"예시답안 {result['created_examples']}개 생성, "
            f"기존 예시 {result['replaced_examples']}개 교체"
        )
    else:
        result = apply_answer_zip_preview(program, batch.preview_data)
        batch.message = (
            f"챕터 {result['created_chapters']}개 생성, "
            f"아이템 {result['created_items']}개 생성, {result['updated_items']}개 업데이트"
        )
    batch.status = AnswerZipImportBatch.STATUS_APPLIED
    batch.applied_at = timezone.now()
    batch.save(update_fields=["status", "applied_at", "message"])
    messages.success(request, f"answer.zip 적용 완료: {batch.message}")
    return redirect("chapter_manage", program_id=program.id)


@login_required
@user_passes_test(is_admin)
def olympiad_example_add(request, item_id):
    item = get_object_or_404(Item, id=item_id, item_type="olympiad")
    
    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url and url_has_allowed_host_and_scheme(url=next_url, allowed_hosts={request.get_host()}):
        redirect_url = next_url
    else:
        redirect_url = reverse("chapter_manage", args=[item.chapter.program.id])

    if request.method != "POST":
        return redirect(redirect_url)

    image = request.FILES.get("image")
    if not image:
        messages.error(request, "등록할 예시답안 이미지 파일을 선택해 주세요.")
        return redirect(redirect_url)

    image_error = validate_olympiad_example_image(image)
    if image_error:
        messages.error(request, image_error)
        return redirect(redirect_url)

    next_order = (item.olympiad_examples.order_by("-order").values_list("order", flat=True).first() or 0) + 1
    OlympiadAnswerExample.objects.create(
        item=item,
        image=image,
        caption=(request.POST.get("caption") or "").strip(),
        order=next_order,
    )
    messages.success(request, f"'{item.title}' 예시답안 이미지가 등록되었습니다.")
    return redirect(redirect_url)


# --- 챕터 관리 ---
@login_required
@user_passes_test(is_admin)
def chapter_create(request, program_id):
    program = get_object_or_404(LearningProgram, id=program_id)
    if request.method == "POST":
        form = ChapterForm(request.POST)
        if form.is_valid():
            chapter = form.save(commit=False)
            chapter.program = program
            chapter.save()
            messages.success(request, f"'{chapter.title}' 챕터가 등록되었습니다.")
            return redirect("chapter_manage", program_id=program.id)
    else:
        last_num = Chapter.objects.filter(program=program).count() + 1
        form = ChapterForm(initial={'number': last_num})
    
    return render(request, "courses/chapter_form.html", {
        "form": form,
        "title": f"[{program.name}] 새 챕터 추가",
        "program": program
    })

@login_required
@user_passes_test(is_admin)
def chapter_edit(request, chapter_id):
    chapter = get_object_or_404(Chapter, id=chapter_id)
    if request.method == "POST":
        form = ChapterForm(request.POST, instance=chapter)
        if form.is_valid():
            form.save()
            messages.success(request, f"'{chapter.title}' 챕터가 수정되었습니다.")
            return redirect("chapter_manage", program_id=chapter.program.id)
    else:
        form = ChapterForm(instance=chapter)
    
    return render(request, "courses/chapter_form.html", {
        "form": form,
        "title": "챕터 수정",
        "chapter": chapter,
        "program": chapter.program,
    })

@login_required
@user_passes_test(is_admin)
def chapter_delete(request, chapter_id):
    chapter = get_object_or_404(Chapter, id=chapter_id)
    program_id = chapter.program.id
    if request.method == "POST":
        title = chapter.title
        chapter.delete()
        messages.success(request, f"'{title}' 챕터가 삭제되었습니다.")
        return redirect("chapter_manage", program_id=program_id)
    return render(request, "courses/chapter_confirm_delete.html", {"chapter": chapter})

# --- (8) 아이템 챕터 이동 ---
@login_required
@user_passes_test(is_admin)
def item_move(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    program = item.chapter.program
    if request.method == "POST":
        target_chapter_id = request.POST.get("target_chapter")
        if target_chapter_id:
            target_chapter = get_object_or_404(Chapter, id=target_chapter_id, program=program)
            old_chapter = item.chapter
            item.chapter = target_chapter
            item.save()
            messages.success(request, f"'{item.title}' 아이템이 [{old_chapter.title}] → [{target_chapter.title}]으로 이동되었습니다.")
        return redirect("chapter_manage", program_id=program.id)
    chapters = Chapter.objects.filter(program=program).exclude(id=item.chapter.id)
    return render(request, "courses/item_move.html", {
        "item": item,
        "chapters": chapters,
        "program": program,
    })

# --- (9) 아이템 등록 ---
@login_required
@user_passes_test(is_admin)
def item_create(request, chapter_id):
    chapter = get_object_or_404(Chapter, id=chapter_id)
    if request.method == "POST":
        form = ItemForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.chapter = chapter
            item.save()
            messages.success(request, f"'{item.title}' 아이템이 등록되었습니다.")
            return redirect("chapter_detail", chapter_id=chapter.id)
    else:
        # 마지막 번호 + 1 자동 제안
        last_num = Item.objects.filter(chapter=chapter).count() + 1
        form = ItemForm(initial={'chapter': chapter, 'number': last_num})
    
    return render(request, "courses/item_form.html", {
        "form": form,
        "title": f"[{chapter.title}] 새 아이템 추가",
        "chapter": chapter
    })

# --- (10) 아이템 수정 ---
@login_required
@user_passes_test(is_admin)
def item_edit(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f"'{item.title}' 아이템이 수정되었습니다.")
            return redirect("item_page", item_id=item.id)
    else:
        form = ItemForm(instance=item)
    
    return render(request, "courses/item_form.html", {
        "form": form,
        "title": "아이템 수정",
        "item": item
    })

# --- (11) 아이템 삭제 ---
@login_required
@user_passes_test(is_admin)
def item_delete(request, item_id):
    item = get_object_or_404(Item, id=item_id)
    chapter_id = item.chapter.id
    if request.method == "POST":
        title = item.title
        item.delete()
        messages.success(request, f"'{title}' 아이템이 삭제되었습니다.")
        return redirect("chapter_detail", chapter_id=chapter_id)
    return render(request, "courses/item_confirm_delete.html", {"item": item})
# --- (8) 학생용 숙제 관리 (홈플레이) ---

@login_required
@require_full_member
def student_homework_list(request):
    selected_member_id = (request.GET.get('member') or '').strip()

    if request.user.is_staff or request.user.is_superuser:
        assignments = HomeworkAssignment.objects.filter(
            is_active=True
        ).select_related('program').prefetch_related(
            'linked_items', 'assigned_users', 'attachments', 'submissions__student'
        ).distinct().order_by('-created_at', '-due_date')

        raw_member_choices = list(User.objects.filter(
            assigned_homeworks__isnull=False
        ).distinct().order_by('username'))

        base_name_counts = Counter(get_member_filter_name(member) for member in raw_member_choices)
        member_choices = []
        for member in raw_member_choices:
            base_name = get_member_filter_name(member)
            has_duplicate_name = base_name_counts[base_name] > 1
            pending_exists = HomeworkAssignment.objects.filter(
                is_active=True,
                assigned_users=member,
            ).exclude(
                submissions__student=member,
                submissions__status=HomeworkSubmission.STATUS_COMPLETED,
            ).exists()
            member_choices.append({
                'id': member.id,
                'label': member.username if has_duplicate_name else base_name,
                'has_pending': pending_exists,
            })

        if selected_member_id.isdigit():
            assignments = assignments.filter(assigned_users__id=int(selected_member_id)).distinct()
    else:
        enrolled_ids = LearningEnrollment.objects.filter(user=request.user).values_list("program_id", flat=True)
        assignments = HomeworkAssignment.objects.filter(
            Q(program_id__in=enrolled_ids, assigned_users__isnull=True) |
            Q(assigned_users=request.user),
            is_active=True
        ).select_related('program').prefetch_related(
            'linked_items', 'assigned_users', 'attachments'
        ).distinct().order_by('-created_at', '-due_date')
        member_choices = []

    submissions_by_assignment = {}
    if not (request.user.is_staff or request.user.is_superuser):
        submissions_by_assignment = {
            submission.assignment_id: submission
            for submission in HomeworkSubmission.objects.filter(
                student=request.user,
                assignment__in=assignments,
            )
        }

    for assignment in assignments:
        items = assignment.linked_items.all()
        if items.exists():
            if request.user.is_staff or request.user.is_superuser:
                completed_count = 0
            else:
                completed_count = UserProgress.objects.filter(
                    user=request.user,
                    item__in=items,
                    completed=True,
                ).count()
            assignment.is_completed = (completed_count == items.count())
            assignment.progress_text = f"{completed_count}/{items.count()}"
        else:
            assignment.is_completed = False
            assignment.progress_text = '제출형 숙제'

        assignment.attachments_list = list(assignment.attachments.all())

        if request.user.is_staff or request.user.is_superuser:
            submissions = list(assignment.submissions.all().order_by('-updated_at'))
            assignment.admin_submissions = submissions
            assigned_users = list(assignment.assigned_users.all().order_by('username'))
            submission_map = {submission.student_id: submission for submission in submissions}
            assignment.admin_member_rows = [
                {
                    'student': assigned_user,
                    'submission': submission_map.get(assigned_user.id),
                }
                for assigned_user in assigned_users
            ]
            assignment.submission = None
            assignment.submission_form = None
            assignment.submission_status_label = ''
            if submissions:
                if any(sub.status == HomeworkSubmission.STATUS_SUBMITTED for sub in submissions):
                    assignment.progress_text = f'제출 {len(submissions)}건'
                elif any(sub.status == HomeworkSubmission.STATUS_REVISION for sub in submissions):
                    assignment.progress_text = f'수정 요청 {len(submissions)}건'
                elif all(sub.status == HomeworkSubmission.STATUS_COMPLETED for sub in submissions):
                    assignment.is_completed = True
                    assignment.progress_text = f'최종 완료 {len(submissions)}건'
            else:
                assignment.progress_text = '제출 대기'
        else:
            submission = submissions_by_assignment.get(assignment.id)
            assignment.submission = submission
            assignment.submission_form = HomeworkSubmissionForm(instance=submission)
            assignment.submission_status_label = submission.get_status_display() if submission else '미제출'
            if submission and submission.status == HomeworkSubmission.STATUS_COMPLETED:
                assignment.is_completed = True
                assignment.progress_text = '평가 완료'
            elif submission and submission.status == HomeworkSubmission.STATUS_SUBMITTED:
                assignment.progress_text = '제출 완료'
            elif submission and submission.status == HomeworkSubmission.STATUS_REVISION:
                assignment.progress_text = '보완 필요'

    return render(request, 'courses/student_homework_list.html', {
        'assignments': assignments,
        'member_choices': member_choices,
        'selected_member_id': selected_member_id,
    })


@login_required
@require_full_member
def student_homework_submit(request, assignment_id):
    if request.user.is_staff or request.user.is_superuser:
        messages.warning(request, '관리자는 학생 제출 화면에서 직접 제출할 수 없습니다.')
        return redirect('student_homework_list')

    assignment = get_object_or_404(HomeworkAssignment, id=assignment_id, is_active=True)
    submission = HomeworkSubmission.objects.filter(assignment=assignment, student=request.user).first()

    if request.method != 'POST':
        return redirect('student_homework_list')

    form = HomeworkSubmissionForm(request.POST, request.FILES, instance=submission)
    if form.is_valid():
        homework_submission = form.save(commit=False)
        homework_submission.assignment = assignment
        homework_submission.student = request.user
        homework_submission.status = HomeworkSubmission.STATUS_SUBMITTED
        homework_submission.reviewed_at = None
        homework_submission.save()
        messages.success(request, f"'{assignment.title}' 숙제를 제출했습니다.")
    else:
        messages.error(request, '제출 내용을 다시 확인해 주세요.')
    return redirect('student_homework_list')


@login_required
@user_passes_test(is_admin)
def homework_submission_action(request, submission_id):
    submission = get_object_or_404(
        HomeworkSubmission.objects.select_related('assignment', 'student'),
        id=submission_id,
    )
    if request.method != 'POST':
        return redirect('student_homework_list')

    action = request.POST.get('action')
    teacher_comment = (request.POST.get('teacher_comment') or '').strip()

    if action == 'revision':
        submission.status = HomeworkSubmission.STATUS_REVISION
        submission.teacher_comment = teacher_comment
        submission.reviewed_at = timezone.now()
        submission.save(update_fields=['status', 'teacher_comment', 'reviewed_at', 'updated_at'])
        messages.success(request, f'{submission.student.username} 학생에게 수정 요청을 보냈습니다.')
    elif action == 'completed':
        if submission.file:
            submission.file.delete(save=False)
        submission.status = HomeworkSubmission.STATUS_COMPLETED
        submission.teacher_comment = teacher_comment
        submission.reviewed_at = timezone.now()
        submission.save(update_fields=['status', 'teacher_comment', 'reviewed_at', 'updated_at'])
        evaluate_homework_badges(submission.student)
        messages.success(request, f'{submission.student.username} 학생 숙제를 최종 완료 처리했습니다.')
    else:
        messages.warning(request, '처리할 작업을 다시 선택해 주세요.')

    return redirect('student_homework_list')


@login_required
@user_passes_test(is_admin)
def homework_assignment_complete(request, assignment_id, student_id):
    assignment = get_object_or_404(HomeworkAssignment.objects.prefetch_related('assigned_users'), id=assignment_id)
    student = get_object_or_404(User, id=student_id)

    if request.method != 'POST':
        return redirect('student_homework_list')

    if not assignment.assigned_users.filter(id=student.id).exists():
        messages.warning(request, '해당 학생은 이 과제의 배정 대상이 아닙니다.')
        return redirect('student_homework_list')

    teacher_comment = (request.POST.get('teacher_comment') or '').strip()
    submission, _ = HomeworkSubmission.objects.get_or_create(
        assignment=assignment,
        student=student,
        defaults={
            'note': '',
        },
    )
    if submission.file:
        submission.file.delete(save=False)
    submission.status = HomeworkSubmission.STATUS_COMPLETED
    submission.teacher_comment = teacher_comment
    submission.reviewed_at = timezone.now()
    submission.save(update_fields=['status', 'teacher_comment', 'reviewed_at', 'updated_at'])
    evaluate_homework_badges(student)
    messages.success(request, f'{student.username} 학생 숙제를 제출 파일 없이 완료 처리했습니다.')
    return redirect('student_homework_list')


@login_required
@user_passes_test(is_admin)
def homework_admin_list(request):
    assignments = HomeworkAssignment.objects.all().select_related('program').order_by('-created_at')
    return render(request, 'courses/homework_admin_list.html', {'assignments': assignments})


@login_required
@user_passes_test(is_admin)
def homework_create(request):
    if request.method == 'POST':
        form = HomeworkForm(request.POST, request.FILES)
        if form.is_valid():
            assignment = form.save()
            for uploaded_file in request.FILES.getlist('attachment_files'):
                HomeworkAttachment.objects.create(
                    assignment=assignment,
                    title=uploaded_file.name,
                    file=uploaded_file,
                )
            messages.success(request, '새 숙제를 등록했습니다.')
            return redirect('homework_admin_list')
    else:
        form = HomeworkForm()
    return render(request, 'courses/homework_form.html', {
        'form': form,
        'title': '숙제 등록',
        'attachments': [],
        'submissions': [],
    })


@login_required
@user_passes_test(is_admin)
def homework_edit(request, homework_id):
    assignment = get_object_or_404(HomeworkAssignment, id=homework_id)
    if request.method == 'POST':
        form = HomeworkForm(request.POST, request.FILES, instance=assignment)
        if form.is_valid():
            form.save()
            delete_ids = request.POST.getlist('delete_attachment_ids')
            if delete_ids:
                HomeworkAttachment.objects.filter(assignment=assignment, id__in=delete_ids).delete()
            for uploaded_file in request.FILES.getlist('attachment_files'):
                HomeworkAttachment.objects.create(
                    assignment=assignment,
                    title=uploaded_file.name,
                    file=uploaded_file,
                )
            messages.success(request, f"'{assignment.title}' 숙제를 수정했습니다.")
            return redirect('homework_edit', homework_id=assignment.id)
    else:
        form = HomeworkForm(instance=assignment)

    submissions = assignment.submissions.select_related('student').order_by('-updated_at')
    return render(request, 'courses/homework_form.html', {
        'form': form,
        'title': '숙제 수정',
        'assignment': assignment,
        'attachments': assignment.attachments.all(),
        'submissions': submissions,
    })


@login_required
@user_passes_test(is_admin)
def homework_delete(request, homework_id):
    assignment = get_object_or_404(HomeworkAssignment, id=homework_id)
    if request.method == 'POST':
        assignment.delete()
        messages.success(request, '숙제를 삭제했습니다.')
        return redirect('homework_admin_list')
    return render(request, 'courses/homework_confirm_delete.html', {'assignment': assignment})


@login_required
@user_passes_test(is_admin)
def homework_submission_review(request, submission_id):
    submission = get_object_or_404(
        HomeworkSubmission.objects.select_related('assignment', 'student', 'assignment__program'),
        id=submission_id,
    )
    if request.method == 'POST':
        form = HomeworkSubmissionReviewForm(request.POST, instance=submission)
        if form.is_valid():
            reviewed_submission = form.save(commit=False)
            reviewed_submission.reviewed_at = timezone.now()
            reviewed_submission.save()
            if reviewed_submission.status == HomeworkSubmission.STATUS_COMPLETED:
                evaluate_homework_badges(reviewed_submission.student)
            messages.success(request, f'{submission.student.username} 학생 제출물을 평가했습니다.')
            return redirect('homework_edit', homework_id=submission.assignment_id)
    else:
        form = HomeworkSubmissionReviewForm(instance=submission)

    return render(request, 'courses/homework_submission_review.html', {
        'form': form,
        'submission': submission,
    })

# --- (10) API: 프로그램 구조 조회 (과정 → 단원 → 문제) ---
@login_required
@user_passes_test(is_admin)
def api_program_structure(request, program_id):
    """과정 ID를 받아 단원(Chapter)과 하위 문제(Item)들을 JSON으로 반환"""
    chapters = Chapter.objects.filter(program_id=program_id).order_by('number')
    data = []
    for ch in chapters:
        items = Item.objects.filter(chapter=ch).order_by('number')
        data.append({
            'chapter_id': ch.id,
            'chapter_title': f"{ch.number}장: {ch.title}",
            'items': [{'id': item.id, 'title': f"{item.number}. {item.title}", 'type': item.get_item_type_display()} for item in items]
        })
    return JsonResponse({'chapters': data})

# --- (11) API: 학생 검색 ---
@login_required
@user_passes_test(is_admin)
def api_search_users(request):
    """학생 이름/아이디로 검색하여 JSON 반환"""
    query = request.GET.get('q', '').strip()
    from django.contrib.auth.models import User
    users = User.objects.filter(is_staff=False, is_superuser=False)
    if query:
        users = users.filter(Q(username__icontains=query) | Q(first_name__icontains=query) | Q(last_name__icontains=query))
    users = users[:50]
    data = [{'id': u.id, 'username': u.username, 'name': u.get_full_name() or u.username} for u in users]
    return JsonResponse({'users': data})

# --- (12) API: 챕터 순서 재배열 ---
@login_required
@user_passes_test(is_admin)
def api_chapters_reorder(request, program_id):
    """드래그앤드롭으로 변경된 챕터 순서 저장"""
    if request.method != "POST":
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        chapter_ids = data.get('chapter_ids', [])
        for index, cid in enumerate(chapter_ids):
            Chapter.objects.filter(id=cid, program_id=program_id).update(number=index + 1)
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# --- (13) API: 아이템 일괄 이동 ---
@login_required
@user_passes_test(is_admin)
def api_items_batch_move(request):
    """선택한 아이템들을 특정 챕터로 일괄 이동"""
    if request.method != "POST":
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        item_ids = data.get('item_ids', [])
        target_chapter_id = data.get('target_chapter_id')
        if not item_ids or not target_chapter_id:
            return JsonResponse({'error': 'item_ids and target_chapter_id required'}, status=400)
        target_chapter = get_object_or_404(Chapter, id=target_chapter_id)
        items = Item.objects.filter(id__in=item_ids)
        count = items.count()
        items.update(chapter=target_chapter)
        # 번호 재정렬
        all_items = Item.objects.filter(chapter=target_chapter).order_by('number')
        for idx, item in enumerate(all_items):
            Item.objects.filter(id=item.id).update(number=idx + 1)
        return JsonResponse({'status': 'ok', 'moved': count})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# --- (14) API: 아이템 순서 재배열 ---
@login_required
@user_passes_test(is_admin)
def api_items_reorder(request, program_id=None):
    """드래그앤드롭으로 변경된 아이템 순서 저장"""
    if request.method != "POST":
        return JsonResponse({'error': 'POST required'}, status=405)
    try:
        data = json.loads(request.body)
        chapter_id = data.get('chapter_id')
        item_ids = data.get('item_ids', [])
        if not chapter_id or not item_ids:
            return JsonResponse({'error': 'chapter_id and item_ids required'}, status=400)
        for index, iid in enumerate(item_ids):
            Item.objects.filter(id=iid, chapter_id=chapter_id).update(number=index + 1)
        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
